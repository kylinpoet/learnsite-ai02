from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.api.deps.auth import require_admin, require_staff
from app.api.deps.db import get_db
from app.models import CurriculumBook, CurriculumLesson, CurriculumUnit, User
from app.schemas.common import ApiResponse

router = APIRouter()
CURRICULUM_IMPORT_REQUIRED_FIELDS = (
    "book_name",
    "subject",
    "edition",
    "grade_scope",
    "term_no",
    "unit_no",
    "unit_title",
    "lesson_no",
    "lesson_title",
)
CURRICULUM_IMPORT_HEADER_LABELS = {
    "book_name": "教材名称",
    "subject": "学科",
    "edition": "版本",
    "grade_scope": "适用范围",
    "term_no": "学期",
    "unit_no": "单元序号",
    "unit_title": "单元标题",
    "lesson_no": "课次序号",
    "lesson_title": "课次标题",
    "lesson_summary": "课次摘要",
}
CURRICULUM_IMPORT_HEADER_ALIASES = {
    "book_name": {"book_name", "book", "教材", "教材名称", "册", "册名", "书名"},
    "subject": {"subject", "学科", "课程", "科目"},
    "edition": {"edition", "版本", "教材版本"},
    "grade_scope": {"grade_scope", "grade", "适用范围", "年级范围", "年级"},
    "term_no": {"term_no", "term", "学期", "学期序号"},
    "unit_no": {"unit_no", "unit", "单元序号", "单元号", "单元"},
    "unit_title": {"unit_title", "unit_name", "单元标题", "单元名称"},
    "lesson_no": {"lesson_no", "lesson", "课次序号", "课程序号", "课次", "课号"},
    "lesson_title": {"lesson_title", "lesson_name", "课次标题", "课题", "课次名称"},
    "lesson_summary": {"lesson_summary", "summary", "课次摘要", "摘要", "说明"},
}


class CurriculumBookPayload(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    subject: str = Field(min_length=1, max_length=50)
    edition: str = Field(min_length=1, max_length=100)
    grade_scope: str = Field(min_length=1, max_length=50)


class CurriculumUnitPayload(BaseModel):
    book_id: int = Field(ge=1)
    term_no: int = Field(ge=1, le=4)
    unit_no: int = Field(ge=1, le=99)
    title: str = Field(min_length=1, max_length=100)


class CurriculumLessonPayload(BaseModel):
    unit_id: int = Field(ge=1)
    lesson_no: int = Field(ge=1, le=99)
    title: str = Field(min_length=1, max_length=120)
    summary: str | None = Field(default=None, max_length=5000)


def normalize_import_header(value: str) -> str:
    return "".join(
        ch
        for ch in value.strip().lower()
        if ch
        not in {" ", "\t", "\r", "\n", "-", "_", "(", ")", "[", "]", "{", "}", "（", "）"}
    )


NORMALIZED_CURRICULUM_IMPORT_ALIASES = {
    field_name: {normalize_import_header(alias) for alias in aliases}
    for field_name, aliases in CURRICULUM_IMPORT_HEADER_ALIASES.items()
}


def decode_tabular_file(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件编码无法识别，请保存为 UTF-8 或 Excel 默认编码后重试")


def load_csv_rows(file_bytes: bytes) -> list[list[object]]:
    text = decode_tabular_file(file_bytes)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader if any(str(cell).strip() for cell in row)]


def load_xlsx_rows(file_bytes: bytes) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="后端缺少 openpyxl，暂时无法导入 .xlsx 文件") from exc

    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(str(cell).strip() for cell in row if cell is not None)
        ]
    finally:
        workbook.close()


def load_import_rows(file_name: str, file_bytes: bytes) -> list[list[object]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_rows(file_bytes)
    if suffix in {".csv", ".txt", ".tsv"}:
        return load_csv_rows(file_bytes)
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="只支持导入 .csv、.txt、.tsv 或 .xlsx 文件")


def resolve_curriculum_import_headers(header_row: list[object]) -> dict[str, int]:
    normalized_headers = [normalize_import_header(str(cell or "")) for cell in header_row]
    header_map: dict[str, int] = {}

    for field_name, aliases in NORMALIZED_CURRICULUM_IMPORT_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                header_map[field_name] = index
                break

    missing = [
        CURRICULUM_IMPORT_HEADER_LABELS[field_name]
        for field_name in CURRICULUM_IMPORT_REQUIRED_FIELDS
        if field_name not in header_map
    ]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"导入表头缺少必要字段：{', '.join(missing)}",
        )
    return header_map


def get_row_value(row: list[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def parse_import_int(
    value: object,
    *,
    row_number: int,
    field_label: str,
    min_value: int,
    max_value: int,
) -> int:
    if isinstance(value, bool):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}格式不正确")

    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数")
        parsed = int(value)
    else:
        text = str(value or "").strip()
        if not text:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行缺少{field_label}")
        try:
            numeric = float(text)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数") from exc
        if not numeric.is_integer():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数")
        parsed = int(numeric)

    if parsed < min_value or parsed > max_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}必须在 {min_value} 到 {max_value} 之间",
        )
    return parsed


def parse_import_text(
    value: object,
    *,
    row_number: int,
    field_label: str,
    max_length: int,
    allow_empty: bool = False,
) -> str | None:
    text = str(value or "").strip()
    if not text:
        if allow_empty:
            return None
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行缺少{field_label}")
    if len(text) > max_length:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}长度不能超过 {max_length} 个字符",
        )
    return text


def row_is_comment(row: list[object]) -> bool:
    for cell in row:
        text = str(cell or "").strip()
        if text:
            return text.startswith("#")
    return False


def build_curriculum_import_rows(rows: list[list[object]]) -> list[dict]:
    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有找到教材数据")

    header_map = resolve_curriculum_import_headers(rows[0])
    parsed_rows: list[dict] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(str(cell or "").strip() for cell in row):
            continue
        if row_is_comment(row):
            continue

        parsed_rows.append(
            {
                "book_name": parse_import_text(
                    get_row_value(row, header_map.get("book_name")),
                    row_number=row_number,
                    field_label="教材名称",
                    max_length=100,
                ),
                "subject": parse_import_text(
                    get_row_value(row, header_map.get("subject")),
                    row_number=row_number,
                    field_label="学科",
                    max_length=50,
                ),
                "edition": parse_import_text(
                    get_row_value(row, header_map.get("edition")),
                    row_number=row_number,
                    field_label="版本",
                    max_length=100,
                ),
                "grade_scope": parse_import_text(
                    get_row_value(row, header_map.get("grade_scope")),
                    row_number=row_number,
                    field_label="适用范围",
                    max_length=50,
                ),
                "term_no": parse_import_int(
                    get_row_value(row, header_map.get("term_no")),
                    row_number=row_number,
                    field_label="学期",
                    min_value=1,
                    max_value=4,
                ),
                "unit_no": parse_import_int(
                    get_row_value(row, header_map.get("unit_no")),
                    row_number=row_number,
                    field_label="单元序号",
                    min_value=1,
                    max_value=99,
                ),
                "unit_title": parse_import_text(
                    get_row_value(row, header_map.get("unit_title")),
                    row_number=row_number,
                    field_label="单元标题",
                    max_length=100,
                ),
                "lesson_no": parse_import_int(
                    get_row_value(row, header_map.get("lesson_no")),
                    row_number=row_number,
                    field_label="课次序号",
                    min_value=1,
                    max_value=99,
                ),
                "lesson_title": parse_import_text(
                    get_row_value(row, header_map.get("lesson_title")),
                    row_number=row_number,
                    field_label="课次标题",
                    max_length=120,
                ),
                "lesson_summary": parse_import_text(
                    get_row_value(row, header_map.get("lesson_summary")),
                    row_number=row_number,
                    field_label="课次摘要",
                    max_length=5000,
                    allow_empty=True,
                ),
            }
        )

    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有找到可用教材数据行")
    return parsed_rows


def import_curriculum_rows(db: Session, parsed_rows: list[dict]) -> dict:
    books = (
        db.scalars(
            select(CurriculumBook).options(
                selectinload(CurriculumBook.units).selectinload(CurriculumUnit.lessons)
            )
        )
        .unique()
        .all()
    )

    book_map: dict[tuple[str, str, str, str], CurriculumBook] = {}
    unit_map: dict[tuple[int, int, int], CurriculumUnit] = {}
    lesson_map: dict[tuple[int, int], CurriculumLesson] = {}

    for book in books:
        book_key = (book.name.strip(), book.subject.strip(), book.edition.strip(), book.grade_scope.strip())
        book_map[book_key] = book
        for unit in book.units:
            unit_map[(book.id, unit.term_no, unit.unit_no)] = unit
            for lesson in unit.lessons:
                lesson_map[(unit.id, lesson.lesson_no)] = lesson

    created_book_count = 0
    created_unit_count = 0
    updated_unit_count = 0
    created_lesson_count = 0
    updated_lesson_count = 0

    for row in parsed_rows:
        book_key = (
            str(row["book_name"]).strip(),
            str(row["subject"]).strip(),
            str(row["edition"]).strip(),
            str(row["grade_scope"]).strip(),
        )
        book = book_map.get(book_key)
        if book is None:
            book = CurriculumBook(
                name=book_key[0],
                subject=book_key[1],
                edition=book_key[2],
                grade_scope=book_key[3],
            )
            db.add(book)
            db.flush()
            book_map[book_key] = book
            created_book_count += 1

        unit_key = (book.id, int(row["term_no"]), int(row["unit_no"]))
        unit = unit_map.get(unit_key)
        unit_title = str(row["unit_title"]).strip()
        if unit is None:
            unit = CurriculumUnit(
                book_id=book.id,
                term_no=unit_key[1],
                unit_no=unit_key[2],
                title=unit_title,
            )
            db.add(unit)
            db.flush()
            unit_map[unit_key] = unit
            created_unit_count += 1
        elif unit.title != unit_title:
            unit.title = unit_title
            updated_unit_count += 1

        lesson_key = (unit.id, int(row["lesson_no"]))
        lesson = lesson_map.get(lesson_key)
        lesson_title = str(row["lesson_title"]).strip()
        lesson_summary = row["lesson_summary"]
        summary_text = str(lesson_summary).strip() if isinstance(lesson_summary, str) else None
        if summary_text == "":
            summary_text = None
        if lesson is None:
            lesson = CurriculumLesson(
                unit_id=unit.id,
                lesson_no=lesson_key[1],
                title=lesson_title,
                summary=summary_text,
            )
            db.add(lesson)
            db.flush()
            lesson_map[lesson_key] = lesson
            created_lesson_count += 1
        else:
            changed = False
            if lesson.title != lesson_title:
                lesson.title = lesson_title
                changed = True
            if lesson.summary != summary_text:
                lesson.summary = summary_text
                changed = True
            if changed:
                updated_lesson_count += 1

    return {
        "processed_row_count": len(parsed_rows),
        "created_book_count": created_book_count,
        "created_unit_count": created_unit_count,
        "updated_unit_count": updated_unit_count,
        "created_lesson_count": created_lesson_count,
        "updated_lesson_count": updated_lesson_count,
    }


def load_books(db: Session) -> list[CurriculumBook]:
    return (
        db.scalars(
            select(CurriculumBook)
            .options(
                selectinload(CurriculumBook.units)
                .selectinload(CurriculumUnit.lessons)
                .selectinload(CurriculumLesson.lesson_plans)
            )
            .order_by(CurriculumBook.id)
        )
        .unique()
        .all()
    )


def serialize_book(book: CurriculumBook) -> dict:
    return {
        "id": book.id,
        "name": book.name,
        "subject": book.subject,
        "edition": book.edition,
        "grade_scope": book.grade_scope,
        "unit_count": len(book.units),
        "lesson_count": sum(len(unit.lessons) for unit in book.units),
        "plan_count": sum(len(lesson.lesson_plans) for unit in book.units for lesson in unit.lessons),
        "units": [
            {
                "id": unit.id,
                "title": unit.title,
                "term_no": unit.term_no,
                "unit_no": unit.unit_no,
                "lesson_count": len(unit.lessons),
                "lessons": [
                    {
                        "id": lesson.id,
                        "title": lesson.title,
                        "lesson_no": lesson.lesson_no,
                        "summary": lesson.summary,
                        "plan_count": len(lesson.lesson_plans),
                        "latest_plan": (
                            {
                                "id": latest_plan.id,
                                "title": latest_plan.title,
                                "assigned_date": latest_plan.assigned_date.isoformat(),
                            }
                            if (
                                latest_plan := max(
                                    lesson.lesson_plans,
                                    key=lambda item: item.assigned_date,
                                    default=None,
                                )
                            )
                            else None
                        ),
                    }
                    for lesson in sorted(unit.lessons, key=lambda item: item.lesson_no)
                ],
            }
            for unit in sorted(book.units, key=lambda item: (item.term_no, item.unit_no))
        ],
    }


def curriculum_tree_payload(db: Session) -> dict:
    books = load_books(db)
    return {"books": [serialize_book(book) for book in books]}


@router.get("/tree", response_model=ApiResponse)
def curriculum_tree(
    _: User = Depends(require_staff),
    db: Session = Depends(get_db),
) -> ApiResponse:
    return ApiResponse(data=curriculum_tree_payload(db))


@router.post("/import", response_model=ApiResponse)
async def import_curriculum_tree(
    file: UploadFile | None = File(default=None),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if file is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要导入的教材文件")
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要导入的教材文件")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件内容为空")

    rows = load_import_rows(file.filename, file_bytes)
    parsed_rows = build_curriculum_import_rows(rows)
    result = import_curriculum_rows(db, parsed_rows)
    db.commit()
    return ApiResponse(
        message=(
            f"教材导入完成：处理 {result['processed_row_count']} 行，"
            f"新增教材 {result['created_book_count']} 本，新增单元 {result['created_unit_count']} 个，"
            f"新增课次 {result['created_lesson_count']} 个。"
        ),
        data={
            **curriculum_tree_payload(db),
            "import_result": result,
        },
    )


@router.post("/books", response_model=ApiResponse)
def create_book(
    payload: CurriculumBookPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    book = CurriculumBook(
        name=payload.name.strip(),
        subject=payload.subject.strip(),
        edition=payload.edition.strip(),
        grade_scope=payload.grade_scope.strip(),
    )
    db.add(book)
    db.commit()
    return ApiResponse(message="教材已创建", data=curriculum_tree_payload(db))


@router.put("/books/{book_id}", response_model=ApiResponse)
def update_book(
    book_id: int,
    payload: CurriculumBookPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    book = db.get(CurriculumBook, book_id)
    if book is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教材不存在")

    book.name = payload.name.strip()
    book.subject = payload.subject.strip()
    book.edition = payload.edition.strip()
    book.grade_scope = payload.grade_scope.strip()
    db.commit()
    return ApiResponse(message="教材已更新", data=curriculum_tree_payload(db))


@router.delete("/books/{book_id}", response_model=ApiResponse)
def delete_book(
    book_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    book = db.scalar(
        select(CurriculumBook)
        .where(CurriculumBook.id == book_id)
        .options(selectinload(CurriculumBook.units).selectinload(CurriculumUnit.lessons))
    )
    if book is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教材不存在")
    if book.units:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="请先删除教材下的单元和课次")

    db.delete(book)
    db.commit()
    return ApiResponse(message="教材已删除", data=curriculum_tree_payload(db))


@router.post("/units", response_model=ApiResponse)
def create_unit(
    payload: CurriculumUnitPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if db.get(CurriculumBook, payload.book_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教材不存在")

    unit = CurriculumUnit(
        book_id=payload.book_id,
        term_no=payload.term_no,
        unit_no=payload.unit_no,
        title=payload.title.strip(),
    )
    db.add(unit)
    db.commit()
    return ApiResponse(message="单元已创建", data=curriculum_tree_payload(db))


@router.put("/units/{unit_id}", response_model=ApiResponse)
def update_unit(
    unit_id: int,
    payload: CurriculumUnitPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    unit = db.get(CurriculumUnit, unit_id)
    if unit is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="单元不存在")
    if db.get(CurriculumBook, payload.book_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教材不存在")

    unit.book_id = payload.book_id
    unit.term_no = payload.term_no
    unit.unit_no = payload.unit_no
    unit.title = payload.title.strip()
    db.commit()
    return ApiResponse(message="单元已更新", data=curriculum_tree_payload(db))


@router.delete("/units/{unit_id}", response_model=ApiResponse)
def delete_unit(
    unit_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    unit = db.scalar(
        select(CurriculumUnit)
        .where(CurriculumUnit.id == unit_id)
        .options(selectinload(CurriculumUnit.lessons))
    )
    if unit is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="单元不存在")
    if unit.lessons:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="请先删除单元下的课次")

    db.delete(unit)
    db.commit()
    return ApiResponse(message="单元已删除", data=curriculum_tree_payload(db))


@router.post("/lessons", response_model=ApiResponse)
def create_lesson(
    payload: CurriculumLessonPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if db.get(CurriculumUnit, payload.unit_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="单元不存在")

    lesson = CurriculumLesson(
        unit_id=payload.unit_id,
        lesson_no=payload.lesson_no,
        title=payload.title.strip(),
        summary=(payload.summary or "").strip() or None,
    )
    db.add(lesson)
    db.commit()
    return ApiResponse(message="课次已创建", data=curriculum_tree_payload(db))


@router.put("/lessons/{lesson_id}", response_model=ApiResponse)
def update_lesson(
    lesson_id: int,
    payload: CurriculumLessonPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    lesson = db.get(CurriculumLesson, lesson_id)
    if lesson is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="课次不存在")
    if db.get(CurriculumUnit, payload.unit_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="单元不存在")

    lesson.unit_id = payload.unit_id
    lesson.lesson_no = payload.lesson_no
    lesson.title = payload.title.strip()
    lesson.summary = (payload.summary or "").strip() or None
    db.commit()
    return ApiResponse(message="课次已更新", data=curriculum_tree_payload(db))


@router.delete("/lessons/{lesson_id}", response_model=ApiResponse)
def delete_lesson(
    lesson_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    lesson = db.scalar(
        select(CurriculumLesson)
        .where(CurriculumLesson.id == lesson_id)
        .options(selectinload(CurriculumLesson.lesson_plans))
    )
    if lesson is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="课次不存在")
    if lesson.lesson_plans:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="已有学案绑定该课次，不能直接删除")

    db.delete(lesson)
    db.commit()
    return ApiResponse(message="课次已删除", data=curriculum_tree_payload(db))
