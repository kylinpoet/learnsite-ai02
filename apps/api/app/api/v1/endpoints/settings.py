from __future__ import annotations

import csv
import json
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from urllib import error as urllib_error
from urllib import request as urllib_request

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session, selectinload

from app.api.deps.auth import require_admin
from app.api.deps.db import get_db
from app.core.security import hash_password
from app.models import (
    AIProvider,
    ClassSeatAssignment,
    ComputerRoom,
    ComputerSeat,
    CurriculumBook,
    CurriculumLesson,
    CurriculumUnit,
    SchoolClass,
    StaffProfile,
    StudentGroup,
    StudentProfile,
    SystemSetting,
    TeacherClassAssignment,
    User,
)
from app.schemas.common import ApiResponse
from app.services.remote_request_headers import build_remote_request_headers
from app.services.assistant_settings import (
    read_assistant_prompt_settings,
    read_assistant_runtime_settings,
    write_assistant_prompt_settings,
    write_assistant_runtime_settings,
)
from app.services.system_settings import (
    normalize_theme_code,
    read_archived_class_ids,
    read_class_archive_records,
    read_system_settings as load_system_settings_payload,
    theme_presets_payload,
    write_archived_class_ids,
    write_class_archive_records,
    write_system_settings as persist_system_settings,
)

router = APIRouter()
ROOM_GRID_MAX = 50
SEAT_IMPORT_REQUIRED_FIELDS = ("row_no", "col_no", "seat_label", "ip_address")
SEAT_IMPORT_HEADER_LABELS = {
    "row_no": "行号",
    "col_no": "列号",
    "seat_label": "座位号",
    "ip_address": "IP 地址",
    "hostname": "主机名",
    "is_enabled": "是否启用",
}
SEAT_IMPORT_HEADER_ALIASES = {
    "row_no": {"row_no", "row", "rowno", "行", "行号", "排"},
    "col_no": {"col_no", "col", "column", "colno", "列", "列号"},
    "seat_label": {"seat_label", "seat", "seatlabel", "seatname", "seatno", "座位", "座位号", "座位名称", "座号"},
    "ip_address": {"ip", "ip_address", "ipaddress", "ip地址", "ipaddr", "地址"},
    "hostname": {"hostname", "host", "主机名", "电脑名", "计算机名"},
    "is_enabled": {"is_enabled", "enabled", "enable", "启用", "是否启用", "状态"},
}
TRUTHY_TEXTS = {"", "1", "true", "yes", "y", "on", "是", "启用", "已启用"}
FALSY_TEXTS = {"0", "false", "no", "n", "off", "否", "禁用", "停用", "未启用"}
STUDENT_IMPORT_REQUIRED_FIELDS = ("display_name",)
STUDENT_IMPORT_HEADER_LABELS = {
    "student_no": "学号",
    "username": "账号",
    "display_name": "姓名",
    "grade_no": "年级",
    "class_no": "班号",
    "class_name": "班级",
    "gender": "性别",
    "password": "初始密码",
    "entry_year": "入学年份",
}
STUDENT_IMPORT_HEADER_ALIASES = {
    "student_no": {"student_no", "student", "学号", "学生编号"},
    "username": {"username", "account", "账号", "登录账号"},
    "display_name": {"name", "display_name", "姓名", "学生姓名"},
    "grade_no": {"grade", "grade_no", "年级", "年级号"},
    "class_no": {"class", "class_no", "班号"},
    "class_name": {"class_name", "classname", "班级", "班级名称"},
    "gender": {"gender", "sex", "性别"},
    "password": {"password", "pwd", "初始密码"},
    "entry_year": {"entry_year", "入学年份", "入学年"},
}


class SystemSettingsPayload(BaseModel):
    school_name: str = Field(min_length=1, max_length=120)
    active_grade_nos: list[int] = Field(min_length=1)
    theme_code: str = Field(default="mango-splash", min_length=1, max_length=30)
    student_register_enabled: bool = False
    assistant_enabled: bool = True
    auto_attendance_on_login: bool = True
    group_drive_file_max_count: int = Field(default=50, ge=1, le=500)
    group_drive_single_file_max_mb: int = Field(default=20, ge=1, le=1024)
    group_drive_allowed_extensions: str = Field(default="", max_length=500)
    class_transfer_review_note_presets_text: str = Field(default="", max_length=6000)
    class_transfer_unreview_reason_presets_text: str = Field(default="", max_length=6000)


class AssistantPromptSettingsPayload(BaseModel):
    general_prompt: str = Field(default="", max_length=12000)
    lesson_prompt: str = Field(default="", max_length=12000)


class AssistantRuntimeSettingsPayload(BaseModel):
    temperature: float = Field(default=0.4, ge=0, le=2)
    top_p: float | None = Field(default=None, ge=0, le=1)
    max_tokens: int | None = Field(default=None, ge=1, le=8192)
    presence_penalty: float | None = Field(default=None, ge=-2, le=2)
    frequency_penalty: float | None = Field(default=None, ge=-2, le=2)
    streaming_enabled: bool = True


class AdminClassPayload(BaseModel):
    grade_no: int = Field(ge=1, le=12)
    class_no: int = Field(ge=1, le=99)
    head_teacher_name: str | None = Field(default=None, max_length=100)
    default_room_id: int | None = Field(default=None, ge=1)


class AdminClassBatchItemPayload(AdminClassPayload):
    pass


class AdminClassBatchPayload(BaseModel):
    items: list[AdminClassBatchItemPayload] = Field(default_factory=list, min_length=1, max_length=300)
    overwrite_existing: bool = False


class AdminPromotionPreviewPayload(BaseModel):
    source_class_ids: list[int] = Field(default_factory=list, min_length=1, max_length=300)
    grade_increment: int = Field(default=1, ge=1, le=3)


class AdminPromotionExecutePayload(AdminPromotionPreviewPayload):
    copy_teacher_assignments: bool = True
    archive_source_classes: bool = True


class AdminTeacherPayload(BaseModel):
    username: str = Field(min_length=1, max_length=50)
    display_name: str = Field(min_length=1, max_length=100)
    title: str | None = Field(default=None, max_length=100)
    is_admin: bool = False
    password: str | None = Field(default=None, min_length=6, max_length=50)
    class_ids: list[int] = Field(default_factory=list)

    @field_validator("password", mode="before")
    @classmethod
    def normalize_blank_password(cls, value: str | None) -> str | None:
        if isinstance(value, str) and not value.strip():
            return None
        return value


class AdminRoomPayload(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    row_count: int = Field(ge=1, le=ROOM_GRID_MAX)
    col_count: int = Field(ge=1, le=ROOM_GRID_MAX)
    description: str | None = Field(default=None, max_length=500)
    ip_prefix: str | None = Field(default=None, max_length=40)
    ip_start: int | None = Field(default=None, ge=1, le=250)


class AdminSeatPayload(BaseModel):
    id: int | None = None
    row_no: int = Field(ge=1, le=ROOM_GRID_MAX)
    col_no: int = Field(ge=1, le=ROOM_GRID_MAX)
    seat_label: str = Field(min_length=1, max_length=30)
    ip_address: str = Field(min_length=1, max_length=64)
    hostname: str | None = Field(default=None, max_length=100)
    is_enabled: bool = True


class AdminSeatListPayload(BaseModel):
    row_count: int = Field(ge=1, le=ROOM_GRID_MAX)
    col_count: int = Field(ge=1, le=ROOM_GRID_MAX)
    seats: list[AdminSeatPayload] = Field(default_factory=list)


class AdminAIProviderPayload(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    provider_type: str = Field(default="openai-compatible", min_length=1, max_length=40)
    base_url: str = Field(min_length=1, max_length=255)
    api_key: str | None = Field(default=None, max_length=255)
    model_name: str = Field(min_length=1, max_length=120)
    is_default: bool = False
    is_enabled: bool = True


class AdminAIProviderModelDiscoveryPayload(BaseModel):
    provider_type: str = Field(default="openai-compatible", min_length=1, max_length=40)
    base_url: str = Field(min_length=1, max_length=255)
    api_key: str | None = Field(default=None, max_length=255)
    provider_id: int | None = Field(default=None, ge=1)


def normalize_import_header(value: str) -> str:
    return "".join(
        ch for ch in value.strip().lower() if ch not in {" ", "\t", "\r", "\n", "-", "_", "(", ")", "[", "]", "{", "}", "（", "）"}
    )


NORMALIZED_SEAT_IMPORT_ALIASES = {
    field_name: {normalize_import_header(alias) for alias in aliases}
    for field_name, aliases in SEAT_IMPORT_HEADER_ALIASES.items()
}
NORMALIZED_STUDENT_IMPORT_ALIASES = {
    field_name: {normalize_import_header(alias) for alias in aliases}
    for field_name, aliases in STUDENT_IMPORT_HEADER_ALIASES.items()
}


def decode_tabular_file(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件编码无法识别，请保存为 UTF-8 或 Excel 默认编码后重试")


def load_csv_rows(file_bytes: bytes) -> list[list[object]]:
    text = decode_tabular_file(file_bytes)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader if any(str(cell).strip() for cell in row)]


def load_xlsx_rows(file_bytes: bytes) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="后端缺少 openpyxl，暂时无法导入 .xlsx 文件") from exc

    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(str(cell).strip() for cell in row if cell is not None)
        ]
    finally:
        workbook.close()


def load_import_rows(file_name: str, file_bytes: bytes) -> list[list[object]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_rows(file_bytes)
    if suffix in {".csv", ".txt", ".tsv"}:
        return load_csv_rows(file_bytes)
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="只支持导入 .csv、.txt、.tsv 或 .xlsx 文件")


def resolve_import_headers(header_row: list[object]) -> dict[str, int]:
    normalized_headers = [normalize_import_header(str(cell or "")) for cell in header_row]
    header_map: dict[str, int] = {}

    for field_name, aliases in NORMALIZED_SEAT_IMPORT_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                header_map[field_name] = index
                break

    missing = [SEAT_IMPORT_HEADER_LABELS[field_name] for field_name in SEAT_IMPORT_REQUIRED_FIELDS if field_name not in header_map]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"导入表头缺少必要字段：{', '.join(missing)}",
        )

    return header_map


def get_row_value(row: list[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def parse_import_int(value: object, *, field_label: str, row_number: int) -> int:
    if isinstance(value, bool):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}格式不正确")

    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数")
        parsed = int(value)
    else:
        text = str(value or "").strip()
        if not text:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行缺少{field_label}")
        try:
            numeric = float(text)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数") from exc
        if not numeric.is_integer():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的{field_label}必须是整数")
        parsed = int(numeric)

    if parsed < 1 or parsed > ROOM_GRID_MAX:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}必须在 1 到 {ROOM_GRID_MAX} 之间",
        )
    return parsed


def parse_import_text(value: object, *, field_label: str, row_number: int, allow_empty: bool = False) -> str | None:
    text = str(value or "").strip()
    if text:
        return text
    if allow_empty:
        return None
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行缺少{field_label}")


def parse_import_enabled(value: object, *, row_number: int) -> bool:
    if isinstance(value, bool):
        return value

    text = str(value or "").strip().lower()
    if text in TRUTHY_TEXTS:
        return True
    if text in FALSY_TEXTS:
        return False
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行的是否启用值无法识别")


def build_import_seat_payload(room: ComputerRoom, rows: list[list[object]]) -> AdminSeatListPayload:
    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有找到可用的座位数据")

    header_map = resolve_import_headers(rows[0])
    merged_seats_by_position: dict[tuple[int, int], AdminSeatPayload] = {
        (seat.row_no, seat.col_no): AdminSeatPayload(
            id=seat.id,
            row_no=seat.row_no,
            col_no=seat.col_no,
            seat_label=seat.seat_label,
            ip_address=seat.ip_address,
            hostname=seat.hostname,
            is_enabled=seat.is_enabled,
        )
        for seat in room.seats
    }

    imported_positions: set[tuple[int, int]] = set()
    imported_max_row = room.row_count
    imported_max_col = room.col_count
    has_import_rows = False

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(str(cell or "").strip() for cell in row):
            continue

        row_no = parse_import_int(get_row_value(row, header_map.get("row_no")), field_label="行号", row_number=row_number)
        col_no = parse_import_int(get_row_value(row, header_map.get("col_no")), field_label="列号", row_number=row_number)
        seat_label = parse_import_text(get_row_value(row, header_map.get("seat_label")), field_label="座位号", row_number=row_number)
        ip_address = parse_import_text(get_row_value(row, header_map.get("ip_address")), field_label="IP 地址", row_number=row_number)
        hostname = parse_import_text(
            get_row_value(row, header_map.get("hostname")),
            field_label="主机名",
            row_number=row_number,
            allow_empty=True,
        )
        is_enabled = parse_import_enabled(get_row_value(row, header_map.get("is_enabled")), row_number=row_number)

        position = (row_no, col_no)
        if position in imported_positions:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"导入文件中重复出现座位坐标：{row_no}-{col_no}")

        existing = merged_seats_by_position.get(position)
        merged_seats_by_position[position] = AdminSeatPayload(
            id=existing.id if existing else None,
            row_no=row_no,
            col_no=col_no,
            seat_label=seat_label or "",
            ip_address=ip_address or "",
            hostname=hostname,
            is_enabled=is_enabled,
        )
        imported_positions.add(position)
        imported_max_row = max(imported_max_row, row_no)
        imported_max_col = max(imported_max_col, col_no)
        has_import_rows = True

    if not has_import_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有找到有效的座位数据行")

    merged_seats = sorted(merged_seats_by_position.values(), key=lambda item: (item.row_no, item.col_no, item.id or 0))
    return AdminSeatListPayload(row_count=imported_max_row, col_count=imported_max_col, seats=merged_seats)


def resolve_student_import_headers(header_row: list[object]) -> dict[str, int]:
    normalized_headers = [normalize_import_header(str(cell or "")) for cell in header_row]
    header_map: dict[str, int] = {}

    for field_name, aliases in NORMALIZED_STUDENT_IMPORT_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                header_map[field_name] = index
                break

    missing_required = [
        STUDENT_IMPORT_HEADER_LABELS[field_name]
        for field_name in STUDENT_IMPORT_REQUIRED_FIELDS
        if field_name not in header_map
    ]
    if missing_required:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"学生导入表头缺少必要字段：{', '.join(missing_required)}",
        )
    if "student_no" not in header_map and "username" not in header_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="学生导入表头需要至少包含“学号”或“账号”字段",
        )
    if "class_name" not in header_map and not {"grade_no", "class_no"}.issubset(header_map):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="学生导入表头需要提供“班级”字段，或同时提供“年级”和“班号”字段",
        )
    return header_map


def parse_student_import_optional_int(
    value: object,
    *,
    row_number: int,
    field_label: str,
    min_value: int | None = None,
    max_value: int | None = None,
) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None

    try:
        parsed = int(float(text))
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}必须是整数",
        ) from exc

    if min_value is not None and parsed < min_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}不能小于 {min_value}",
        )
    if max_value is not None and parsed > max_value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行的{field_label}不能大于 {max_value}",
        )
    return parsed


def normalize_student_import_gender(value: str | None) -> str:
    if not value:
        return "未知"
    text = value.strip()
    if not text:
        return "未知"
    alias = {
        "男": "男",
        "male": "男",
        "m": "男",
        "boy": "男",
        "女": "女",
        "female": "女",
        "f": "女",
        "girl": "女",
    }
    return alias.get(text.lower(), text)


def build_student_import_preview(
    rows: list[list[object]],
    db: Session,
    *,
    default_password: str,
    update_existing: bool,
) -> dict:
    if len(rows) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件中没有找到学生数据")

    header_map = resolve_student_import_headers(rows[0])
    archived_class_ids = read_archived_class_ids(db)
    classes = [
        item
        for item in db.scalars(select(SchoolClass).order_by(SchoolClass.grade_no.asc(), SchoolClass.class_no.asc())).all()
        if item.id not in archived_class_ids
    ]
    class_by_name = {item.class_name: item for item in classes}
    class_by_grade_no = {(item.grade_no, item.class_no): item for item in classes}

    users = db.scalars(
        select(User)
        .where(User.user_type == "student")
        .options(selectinload(User.student_profile))
    ).all()
    user_by_username = {item.username: item for item in users}
    profile_by_student_no = {
        item.student_profile.student_no: item.student_profile
        for item in users
        if item.student_profile is not None
    }

    seen_row_keys: set[tuple[str, str]] = set()
    imported_rows: list[dict] = []
    skipped_rows: list[dict] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(str(cell or "").strip() for cell in row):
            continue

        display_name = parse_import_text(
            get_row_value(row, header_map.get("display_name")),
            field_label="姓名",
            row_number=row_number,
        )
        student_no = parse_import_text(
            get_row_value(row, header_map.get("student_no")),
            field_label="学号",
            row_number=row_number,
            allow_empty=True,
        )
        username = parse_import_text(
            get_row_value(row, header_map.get("username")),
            field_label="账号",
            row_number=row_number,
            allow_empty=True,
        )
        resolved_student_no = (student_no or username or "").strip()
        resolved_username = (username or student_no or "").strip()
        if not resolved_student_no or not resolved_username:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_number} 行缺少学号或账号",
            )

        class_name = parse_import_text(
            get_row_value(row, header_map.get("class_name")),
            field_label="班级",
            row_number=row_number,
            allow_empty=True,
        )
        grade_no = parse_student_import_optional_int(
            get_row_value(row, header_map.get("grade_no")),
            row_number=row_number,
            field_label="年级",
            min_value=1,
            max_value=12,
        )
        class_no = parse_student_import_optional_int(
            get_row_value(row, header_map.get("class_no")),
            row_number=row_number,
            field_label="班号",
            min_value=1,
            max_value=99,
        )
        target_class = None
        if class_name:
            target_class = class_by_name.get(class_name)
        elif grade_no is not None and class_no is not None:
            target_class = class_by_grade_no.get((grade_no, class_no))

        if target_class is None:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "student_no": resolved_student_no,
                    "username": resolved_username,
                    "display_name": display_name,
                    "status": "skipped",
                    "reason": "找不到目标班级",
                }
            )
            continue

        gender = normalize_student_import_gender(
            parse_import_text(
                get_row_value(row, header_map.get("gender")),
                field_label="性别",
                row_number=row_number,
                allow_empty=True,
            )
        )
        entry_year = parse_student_import_optional_int(
            get_row_value(row, header_map.get("entry_year")),
            row_number=row_number,
            field_label="入学年份",
            min_value=2000,
            max_value=2100,
        ) or (2026 - (target_class.grade_no - 6))
        row_password = parse_import_text(
            get_row_value(row, header_map.get("password")),
            field_label="初始密码",
            row_number=row_number,
            allow_empty=True,
        )
        password = (row_password or default_password).strip()
        if len(password) < 6:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_number} 行的密码不能少于 6 位",
            )

        row_key = (resolved_student_no, resolved_username)
        if row_key in seen_row_keys:
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "student_no": resolved_student_no,
                    "username": resolved_username,
                    "display_name": display_name,
                    "status": "skipped",
                    "reason": "导入文件中出现重复学号/账号",
                }
            )
            continue
        seen_row_keys.add(row_key)

        existing_profile = profile_by_student_no.get(resolved_student_no)
        existing_user = user_by_username.get(resolved_username)
        if not update_existing and (existing_profile is not None or existing_user is not None):
            skipped_rows.append(
                {
                    "row_number": row_number,
                    "student_no": resolved_student_no,
                    "username": resolved_username,
                    "display_name": display_name,
                    "status": "skipped",
                    "reason": "学号或账号已存在",
                }
            )
            continue

        imported_rows.append(
            {
                "row_number": row_number,
                "student_no": resolved_student_no,
                "username": resolved_username,
                "display_name": display_name,
                "class_id": target_class.id,
                "class_name": target_class.class_name,
                "grade_no": target_class.grade_no,
                "gender": gender,
                "entry_year": entry_year,
                "password": password,
                "existing_profile": existing_profile,
                "existing_user": existing_user,
            }
        )

    created_count = 0
    updated_count = 0
    for row in imported_rows:
        existing_profile = row["existing_profile"]
        existing_user = row["existing_user"]
        if existing_profile is not None:
            user = existing_profile.user or db.get(User, existing_profile.user_id)
            if user is None:
                continue
            user.username = row["username"]
            user.display_name = row["display_name"]
            user.password_hash = hash_password(row["password"])
            user.is_active = True
            existing_profile.student_no = row["student_no"]
            existing_profile.class_id = row["class_id"]
            existing_profile.grade_no = row["grade_no"]
            existing_profile.gender = row["gender"]
            existing_profile.entry_year = row["entry_year"]
            updated_count += 1
            continue

        if existing_user is not None and existing_user.student_profile is not None:
            profile = existing_user.student_profile
            existing_user.username = row["username"]
            existing_user.display_name = row["display_name"]
            existing_user.password_hash = hash_password(row["password"])
            existing_user.is_active = True
            profile.student_no = row["student_no"]
            profile.class_id = row["class_id"]
            profile.grade_no = row["grade_no"]
            profile.gender = row["gender"]
            profile.entry_year = row["entry_year"]
            updated_count += 1
            continue

        user = User(
            username=row["username"],
            password_hash=hash_password(row["password"]),
            display_name=row["display_name"],
            user_type="student",
            is_active=True,
        )
        db.add(user)
        db.flush()
        db.add(
            StudentProfile(
                user_id=user.id,
                student_no=row["student_no"],
                grade_no=row["grade_no"],
                class_id=row["class_id"],
                gender=row["gender"],
                entry_year=row["entry_year"],
            )
        )
        created_count += 1

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": len(skipped_rows),
        "skipped_rows": skipped_rows,
        "processed_count": len(imported_rows) + len(skipped_rows),
    }


def validate_room_seat_payload(room: ComputerRoom, payload: AdminSeatListPayload, db: Session) -> None:
    existing_by_id = {seat.id: seat for seat in room.seats}
    seen_ids: set[int] = set()
    seen_positions: set[tuple[int, int]] = set()
    seen_ips: set[str] = set()

    for item in payload.seats:
        if item.row_no > payload.row_count or item.col_no > payload.col_count:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="存在超出机房网格范围的座位")

        if item.id is not None:
            if item.id not in existing_by_id:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="提交中包含不属于当前机房的座位记录")
            if item.id in seen_ids:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="提交中存在重复的座位记录")
            seen_ids.add(item.id)

        position = (item.row_no, item.col_no)
        if position in seen_positions:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"存在重复座位坐标：{item.row_no}-{item.col_no}")
        seen_positions.add(position)

        ip_address = item.ip_address.strip()
        if not ip_address:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"座位 {item.row_no}-{item.col_no} 缺少 IP 地址")
        normalized_ip = ip_address.lower()
        if normalized_ip in seen_ips:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"存在重复 IP 地址：{ip_address}")
        seen_ips.add(normalized_ip)

    ip_addresses = [item.ip_address.strip() for item in payload.seats]
    if not ip_addresses:
        return

    conflicting_seats = db.scalars(
        select(ComputerSeat).where(
            ComputerSeat.room_id != room.id,
            ComputerSeat.ip_address.in_(ip_addresses),
        )
    ).all()
    if conflicting_seats:
        conflict_ips = sorted({seat.ip_address for seat in conflicting_seats})
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"以下 IP 已被其他机房占用：{', '.join(conflict_ips)}",
        )


def read_system_settings(db: Session) -> dict:
    return load_system_settings_payload(db)


def write_system_settings(payload: SystemSettingsPayload, db: Session) -> None:
    data = payload.model_dump()
    data["theme_code"] = normalize_theme_code(str(data.get("theme_code") or ""))
    persist_system_settings(data, db)


def normalize_ai_provider_value(value: str | None) -> str:
    return (value or "").strip()


def mask_ai_provider_key(api_key: str | None) -> str:
    if not api_key:
        return ""
    if len(api_key) <= 8:
        return "*" * len(api_key)
    return f"{api_key[:4]}{'*' * (len(api_key) - 8)}{api_key[-4:]}"


def serialize_ai_provider(provider: AIProvider) -> dict:
    return {
        "id": provider.id,
        "name": provider.name,
        "provider_type": provider.provider_type,
        "base_url": provider.base_url,
        "model_name": provider.model_name,
        "is_default": provider.is_default,
        "is_enabled": provider.is_enabled,
        "has_api_key": bool(provider.api_key),
        "masked_api_key": mask_ai_provider_key(provider.api_key),
        "updated_at": provider.updated_at.isoformat() if provider.updated_at else None,
    }


def ai_provider_items_payload(db: Session) -> dict:
    providers = db.scalars(select(AIProvider).order_by(AIProvider.is_default.desc(), AIProvider.id.asc())).all()
    return {"items": [serialize_ai_provider(item) for item in providers]}


def candidate_ai_provider_model_urls(base_url: str) -> list[str]:
    normalized = base_url.rstrip("/")
    if not normalized:
        return []

    candidates: list[str] = []

    def append(url: str) -> None:
        cleaned = url.rstrip("/")
        if cleaned and cleaned not in candidates:
            candidates.append(cleaned)

    if normalized.endswith("/models"):
        append(normalized)
    elif normalized.endswith("/chat/completions"):
        append(f"{normalized.rsplit('/chat/completions', maxsplit=1)[0]}/models")
    elif normalized.endswith("/v1"):
        append(f"{normalized}/models")
    else:
        append(f"{normalized}/models")
        append(f"{normalized}/v1/models")

    return candidates


def parse_ai_provider_models_payload(raw_payload: object) -> list[str]:
    if not isinstance(raw_payload, dict):
        raise ValueError("模型服务返回格式不正确")

    data = raw_payload.get("data")
    if not isinstance(data, list):
        raise ValueError("模型服务未返回 data 列表")

    models: list[str] = []
    seen: set[str] = set()
    for item in data:
        if not isinstance(item, dict):
            continue
        model_id = str(item.get("id") or "").strip()
        if not model_id or model_id in seen:
            continue
        seen.add(model_id)
        models.append(model_id)

    if not models:
        raise ValueError("模型服务返回成功，但未解析到可用模型")
    return sorted(models, key=str.lower)


def extract_ai_provider_error_message(body: str) -> str:
    text = body.strip()
    if not text:
        return ""
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return text

    if isinstance(payload, dict):
        error = payload.get("error")
        if isinstance(error, dict):
            message = str(error.get("message") or "").strip()
            if message:
                return message
        message = str(payload.get("message") or "").strip()
        if message:
            return message
    return text


def build_ai_provider_request_headers(api_key: str | None) -> dict[str, str]:
    return build_remote_request_headers(api_key)


def summarize_ai_provider_http_error(status_code: int, provider_message: str) -> str:
    lowered = provider_message.lower()
    if status_code == 401:
        return "鉴权失败，请检查 API Key 是否正确，或当前密钥是否具备读取模型列表的权限"
    if status_code == 403:
        if any(
            token in lowered
            for token in ("cloudflare", "access denied", "blocked access", "browser signature", "owner has blocked")
        ):
            return "访问被服务端拒绝，请检查服务端的安全策略、白名单或防火墙配置"
        return "权限不足，请检查 API Key 是否正确，或当前密钥是否具备读取模型列表的权限"
    if status_code == 404:
        return "未找到模型列表接口，请检查 Base URL 是否正确，或服务是否支持标准 /v1/models"
    if status_code == 408:
        return "请求超时，请稍后重试，或检查服务端响应速度"
    if status_code == 429:
        return "请求过于频繁，当前服务正在限流，请稍后再试"
    if 500 <= status_code < 600:
        return "模型服务暂时不可用，请稍后重试"
    if provider_message:
        return provider_message
    return f"请求失败（HTTP {status_code}）"


def summarize_ai_provider_network_error(error: Exception) -> str:
    text = str(error).strip()
    lowered = text.lower()
    if "timed out" in lowered or "timeout" in lowered:
        return "连接超时，请检查网络连通性，或确认服务地址可从当前服务器访问"
    if "connection refused" in lowered or "actively refused" in lowered:
        return "连接被拒绝，请确认服务地址、端口正确，并且服务已经启动"
    if "getaddrinfo failed" in lowered or "name or service not known" in lowered:
        return "域名解析失败，请检查 Base URL 是否填写正确"
    if "certificate" in lowered or "ssl" in lowered:
        return "HTTPS 证书校验失败，请检查证书配置"
    if text:
        return f"网络请求失败：{text}"
    return "无法连接到模型服务，请检查网络与服务地址"


def summarize_ai_provider_payload_error(message: str) -> str:
    text = message.strip()
    if not text:
        return "接口返回格式不符合预期"
    if "未返回 data 列表" in text:
        return "接口已响应，但返回结构不是标准模型列表格式"
    if "未解析到可用模型" in text:
        return "接口已响应，但暂未读取到可用模型"
    if "返回格式不正确" in text:
        return "接口已响应，但返回格式不正确"
    return text


def build_ai_provider_lookup_detail(attempts: list[str]) -> str:
    if not attempts:
        return (
            "未能自动获取模型列表。"
            "请检查 Base URL、API Key 和服务兼容性；如果该服务不提供标准 /v1/models，"
            "也可以直接手动填写模型名称。"
        )

    summary = "；".join(f"{index + 1}. {item}" for index, item in enumerate(attempts[:2]))
    return (
        "未能自动获取模型列表。"
        "请检查 Base URL 是否正确、API Key 是否具备列出模型的权限，"
        "如果该服务不提供标准 /v1/models，也可以直接手动填写模型名称。"
        f" 已尝试：{summary}"
    )


def fetch_ai_provider_models(
    *,
    provider_type: str,
    base_url: str,
    api_key: str | None,
) -> tuple[list[str], str]:
    if provider_type != "openai-compatible":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前仅支持 OpenAI Compatible 服务自动获取模型")

    urls = candidate_ai_provider_model_urls(base_url)
    if not urls:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请先填写有效的 Base URL")

    headers = build_ai_provider_request_headers(api_key)

    errors: list[str] = []
    for url in urls:
        request = urllib_request.Request(url, headers=headers, method="GET")
        try:
            with urllib_request.urlopen(request, timeout=15) as response:
                raw_payload = json.loads(response.read().decode("utf-8"))
            return parse_ai_provider_models_payload(raw_payload), url
        except urllib_error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="ignore")
            provider_message = extract_ai_provider_error_message(body)
            reason = summarize_ai_provider_http_error(exc.code, provider_message)
            if provider_message and provider_message != reason:
                reason = f"{reason}（服务返回：{provider_message}）"
            errors.append(f"{url} -> {reason}")
        except urllib_error.URLError as exc:
            errors.append(f"{url} -> {summarize_ai_provider_network_error(exc)}")
        except TimeoutError as exc:
            errors.append(f"{url} -> {summarize_ai_provider_network_error(exc)}")
        except json.JSONDecodeError:
            errors.append(f"{url} -> 接口已响应，但返回内容不是合法 JSON，可能不是标准模型列表地址")
        except ValueError as exc:
            errors.append(f"{url} -> {summarize_ai_provider_payload_error(str(exc))}")

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=build_ai_provider_lookup_detail(errors),
    )


def serialize_class(school_class: SchoolClass) -> dict:
    return {
        "id": school_class.id,
        "grade_no": school_class.grade_no,
        "class_no": school_class.class_no,
        "class_name": school_class.class_name,
        "head_teacher_name": school_class.head_teacher_name,
        "default_room_id": school_class.default_room_id,
        "student_count": len(school_class.students),
    }


def serialize_archived_class(school_class: SchoolClass, archive_record: dict | None = None) -> dict:
    return {
        "id": school_class.id,
        "grade_no": school_class.grade_no,
        "class_no": school_class.class_no,
        "class_name": school_class.class_name,
        "head_teacher_name": school_class.head_teacher_name,
        "default_room_id": school_class.default_room_id,
        "student_count": len(school_class.students),
        "original_class_name": str(archive_record.get("source_class_name") or "") if archive_record else "",
        "promoted_to_class_id": (
            int(archive_record["target_class_id"])
            if archive_record and isinstance(archive_record.get("target_class_id"), int)
            else None
        ),
        "promoted_to_class_name": (
            str(archive_record.get("target_class_name") or "")
            if archive_record
            else ""
        ),
        "moved_student_count": (
            int(archive_record["moved_student_count"])
            if archive_record and isinstance(archive_record.get("moved_student_count"), int)
            else 0
        ),
        "archived_at": (
            str(archive_record.get("archived_at") or "")
            if archive_record
            else None
        ),
    }


def serialize_teacher(user: User, *, active_class_ids: set[int]) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "title": user.staff_profile.title if user.staff_profile else None,
        "is_admin": bool(user.staff_profile and user.staff_profile.is_admin),
        "class_ids": sorted(
            assignment.class_id
            for assignment in user.teacher_class_assignments
            if assignment.class_id in active_class_ids
        ),
    }


def serialize_room(room: ComputerRoom) -> dict:
    return {
        "id": room.id,
        "name": room.name,
        "row_count": room.row_count,
        "col_count": room.col_count,
        "description": room.description,
        "seat_count": len(room.seats),
        "seats": [
            {
                "id": seat.id,
                "row_no": seat.row_no,
                "col_no": seat.col_no,
                "seat_label": seat.seat_label,
                "ip_address": seat.ip_address,
                "hostname": seat.hostname,
                "is_enabled": seat.is_enabled,
            }
            for seat in sorted(room.seats, key=lambda item: (item.row_no, item.col_no, item.id))
        ],
    }


def serialize_curriculum_snapshot(db: Session) -> list[dict]:
    books = (
        db.scalars(
            select(CurriculumBook)
            .options(
                selectinload(CurriculumBook.units)
                .selectinload(CurriculumUnit.lessons)
                .selectinload(CurriculumLesson.lesson_plans)
            )
            .order_by(CurriculumBook.id)
        )
        .unique()
        .all()
    )
    return [
        {
            "id": book.id,
            "name": book.name,
            "subject": book.subject,
            "edition": book.edition,
            "grade_scope": book.grade_scope,
            "unit_count": len(book.units),
            "lesson_count": sum(len(unit.lessons) for unit in book.units),
            "plan_count": sum(len(lesson.lesson_plans) for unit in book.units for lesson in unit.lessons),
        }
        for book in books
    ]


def make_class_display_name(grade_no: int, class_no: int) -> str:
    return f"{grade_no}{class_no:02d}班"


def make_archived_class_name(base_name: str, archived_at: datetime, existing_names: set[str]) -> str:
    date_part = archived_at.strftime("%Y%m%d")
    suffix_index = 0
    while True:
        suffix = f"-{suffix_index}" if suffix_index else ""
        candidate = f"{base_name}[归档{date_part}{suffix}]"
        if candidate not in existing_names:
            return candidate
        suffix_index += 1


def build_promotion_preview(
    source_classes: list[SchoolClass],
    active_classes_by_grade_no: dict[tuple[int, int], SchoolClass],
    *,
    grade_increment: int,
) -> dict:
    preview_items: list[dict] = []
    blocked_count = 0
    for school_class in sorted(source_classes, key=lambda item: (item.grade_no, item.class_no, item.id)):
        target_grade_no = school_class.grade_no + grade_increment
        if target_grade_no > 12:
            blocked_count += 1
            preview_items.append(
                {
                    "source_class_id": school_class.id,
                    "source_class_name": school_class.class_name,
                    "source_grade_no": school_class.grade_no,
                    "source_class_no": school_class.class_no,
                    "student_count": len(school_class.students),
                    "target_grade_no": target_grade_no,
                    "target_class_no": school_class.class_no,
                    "target_class_name": make_class_display_name(target_grade_no, school_class.class_no),
                    "target_class_id": None,
                    "target_exists": False,
                    "status": "blocked",
                    "reason": "目标年级超过 12，无法继续升班",
                }
            )
            continue

        target = active_classes_by_grade_no.get((target_grade_no, school_class.class_no))
        preview_items.append(
            {
                "source_class_id": school_class.id,
                "source_class_name": school_class.class_name,
                "source_grade_no": school_class.grade_no,
                "source_class_no": school_class.class_no,
                "student_count": len(school_class.students),
                "target_grade_no": target_grade_no,
                "target_class_no": school_class.class_no,
                "target_class_name": make_class_display_name(target_grade_no, school_class.class_no),
                "target_class_id": target.id if target is not None else None,
                "target_exists": target is not None,
                "status": "ready",
                "reason": "将迁移学生并创建/复用目标班级",
            }
        )

    return {
        "grade_increment": grade_increment,
        "items": preview_items,
        "summary": {
            "source_class_count": len(source_classes),
            "blocked_count": blocked_count,
            "ready_count": len(source_classes) - blocked_count,
            "student_count": sum(len(item.students) for item in source_classes),
        },
    }


def bootstrap_payload(db: Session) -> dict:
    archived_class_ids = read_archived_class_ids(db)
    archive_records = read_class_archive_records(db)
    archive_record_by_class_id = {
        int(item["source_class_id"]): item
        for item in archive_records
        if isinstance(item, dict) and isinstance(item.get("source_class_id"), int)
    }

    all_classes = db.scalars(
        select(SchoolClass)
        .options(selectinload(SchoolClass.students))
        .order_by(SchoolClass.grade_no.asc(), SchoolClass.class_no.asc())
    ).all()
    active_classes = [item for item in all_classes if item.id not in archived_class_ids]
    archived_classes = [item for item in all_classes if item.id in archived_class_ids]
    active_class_id_set = {item.id for item in active_classes}

    teachers = db.scalars(
        select(User)
        .where(User.user_type == "staff")
        .options(
            selectinload(User.staff_profile),
            selectinload(User.teacher_class_assignments),
        )
        .order_by(User.id.asc())
    ).all()
    rooms = (
        db.scalars(
            select(ComputerRoom)
            .options(selectinload(ComputerRoom.seats))
            .order_by(ComputerRoom.id.asc())
        )
        .unique()
        .all()
    )

    return {
        "system": read_system_settings(db),
        "theme_presets": theme_presets_payload(),
        "assistant_prompts": read_assistant_prompt_settings(db),
        "assistant_runtime": read_assistant_runtime_settings(db),
        "classes": [serialize_class(item) for item in active_classes],
        "archived_classes": [
            serialize_archived_class(item, archive_record_by_class_id.get(item.id))
            for item in archived_classes
        ],
        "teachers": [serialize_teacher(item, active_class_ids=active_class_id_set) for item in teachers],
        "rooms": [serialize_room(item) for item in rooms],
        "curriculum_books": serialize_curriculum_snapshot(db),
        "stats": {
            "class_count": len(active_classes),
            "teacher_count": len(teachers),
            "student_count": sum(len(item.students) for item in active_classes),
            "room_count": len(rooms),
            "archived_class_count": len(archived_classes),
        },
    }


def create_room_seats(room: ComputerRoom, ip_prefix: str | None, ip_start: int | None) -> list[ComputerSeat]:
    prefix = (ip_prefix or f"192.168.{room.id}.").strip()
    start = ip_start or 10
    seats: list[ComputerSeat] = []
    seat_index = 0
    for row_no in range(1, room.row_count + 1):
        for col_no in range(1, room.col_count + 1):
            seat_index += 1
            seats.append(
                ComputerSeat(
                    room_id=room.id,
                    row_no=row_no,
                    col_no=col_no,
                    seat_label=f"{row_no}-{col_no}",
                    ip_address=f"{prefix}{start + seat_index - 1}",
                    hostname=f"pc-{room.id}-{seat_index:02d}",
                    is_enabled=True,
                )
            )
    return seats


@router.get("/system", response_model=ApiResponse)
def system_settings(db: Session = Depends(get_db)) -> ApiResponse:
    return ApiResponse(data=read_system_settings(db))


@router.put("/system", response_model=ApiResponse)
def update_system_settings(
    payload: SystemSettingsPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    write_system_settings(payload, db)
    db.commit()
    return ApiResponse(message="系统参数已更新", data=read_system_settings(db))


@router.get("/assistant-prompts", response_model=ApiResponse)
def assistant_prompt_settings(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    return ApiResponse(data=read_assistant_prompt_settings(db))


@router.put("/assistant-prompts", response_model=ApiResponse)
def update_assistant_prompt_settings(
    payload: AssistantPromptSettingsPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    write_assistant_prompt_settings(payload.model_dump(), db)
    db.commit()
    return ApiResponse(
        message="AI companion prompts updated",
        data=read_assistant_prompt_settings(db),
    )


@router.get("/assistant-runtime", response_model=ApiResponse)
def assistant_runtime_settings(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    return ApiResponse(data=read_assistant_runtime_settings(db))


@router.put("/assistant-runtime", response_model=ApiResponse)
def update_assistant_runtime_settings(
    payload: AssistantRuntimeSettingsPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    write_assistant_runtime_settings(payload.model_dump(), db)
    db.commit()
    return ApiResponse(
        message="AI 学伴运行参数已更新",
        data=read_assistant_runtime_settings(db),
    )


@router.get("/admin/bootstrap", response_model=ApiResponse)
def admin_bootstrap(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    return ApiResponse(data=bootstrap_payload(db))


@router.get("/ai-providers", response_model=ApiResponse)
def list_ai_providers(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    return ApiResponse(data=ai_provider_items_payload(db))


@router.post("/ai-providers/discover-models", response_model=ApiResponse)
def discover_ai_provider_models(
    payload: AdminAIProviderModelDiscoveryPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    provider_type = normalize_ai_provider_value(payload.provider_type) or "openai-compatible"
    base_url = normalize_ai_provider_value(payload.base_url).rstrip("/")
    api_key = normalize_ai_provider_value(payload.api_key)

    provider: AIProvider | None = None
    if payload.provider_id is not None:
        provider = db.get(AIProvider, payload.provider_id)
        if provider is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI Provider 不存在")

    effective_api_key = api_key or (provider.api_key if provider is not None else "")
    items, resolved_url = fetch_ai_provider_models(
        provider_type=provider_type,
        base_url=base_url,
        api_key=effective_api_key or None,
    )
    return ApiResponse(
        message=f"已获取 {len(items)} 个模型",
        data={"items": items, "resolved_url": resolved_url},
    )


@router.post("/ai-providers", response_model=ApiResponse)
def create_ai_provider(
    payload: AdminAIProviderPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    name = normalize_ai_provider_value(payload.name)
    provider_type = normalize_ai_provider_value(payload.provider_type) or "openai-compatible"
    base_url = normalize_ai_provider_value(payload.base_url).rstrip("/")
    api_key = normalize_ai_provider_value(payload.api_key)
    model_name = normalize_ai_provider_value(payload.model_name)

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Provider 名称不能为空")
    if not base_url:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Base URL 不能为空")
    if not model_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模型名称不能为空")
    if not api_key:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="新建 AI Provider 时必须提供 API Key")
    if db.scalar(select(AIProvider).where(AIProvider.name == name)) is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="AI Provider 名称已存在")

    provider = AIProvider(
        name=name,
        provider_type=provider_type,
        base_url=base_url,
        api_key=api_key,
        model_name=model_name,
        is_default=payload.is_default,
        is_enabled=payload.is_enabled,
    )
    db.add(provider)
    db.flush()

    providers = db.scalars(select(AIProvider).order_by(AIProvider.id.asc())).all()
    should_make_default = payload.is_default or len(providers) == 1
    if should_make_default:
        for item in providers:
            item.is_default = item.id == provider.id
    elif not any(item.is_default for item in providers):
        provider.is_default = True

    db.commit()
    return ApiResponse(message="AI Provider 已创建", data=ai_provider_items_payload(db))


@router.put("/ai-providers/{provider_id}", response_model=ApiResponse)
def update_ai_provider(
    provider_id: int,
    payload: AdminAIProviderPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    provider = db.get(AIProvider, provider_id)
    if provider is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI Provider 不存在")

    name = normalize_ai_provider_value(payload.name)
    provider_type = normalize_ai_provider_value(payload.provider_type) or "openai-compatible"
    base_url = normalize_ai_provider_value(payload.base_url).rstrip("/")
    api_key = normalize_ai_provider_value(payload.api_key)
    model_name = normalize_ai_provider_value(payload.model_name)

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Provider 名称不能为空")
    if not base_url:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Base URL 不能为空")
    if not model_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="模型名称不能为空")
    duplicate = db.scalar(select(AIProvider).where(AIProvider.name == name, AIProvider.id != provider_id))
    if duplicate is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="AI Provider 名称已存在")

    was_default = provider.is_default
    provider.name = name
    provider.provider_type = provider_type
    provider.base_url = base_url
    if api_key:
        provider.api_key = api_key
    provider.model_name = model_name
    provider.is_default = payload.is_default
    provider.is_enabled = payload.is_enabled

    providers = db.scalars(select(AIProvider).order_by(AIProvider.id.asc())).all()
    if payload.is_default:
        for item in providers:
            item.is_default = item.id == provider.id
    elif was_default:
        fallback = next((item for item in providers if item.id != provider.id), provider)
        for item in providers:
            item.is_default = item.id == fallback.id

    db.commit()
    return ApiResponse(message="AI Provider 已更新", data=ai_provider_items_payload(db))


@router.delete("/ai-providers/{provider_id}", response_model=ApiResponse)
def delete_ai_provider(
    provider_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    provider = db.get(AIProvider, provider_id)
    if provider is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI Provider 不存在")

    was_default = provider.is_default
    db.delete(provider)
    db.flush()

    remaining = db.scalars(select(AIProvider).order_by(AIProvider.id.asc())).all()
    if was_default and remaining:
        remaining[0].is_default = True
        for item in remaining[1:]:
            item.is_default = False

    db.commit()
    return ApiResponse(message="AI Provider 已删除", data=ai_provider_items_payload(db))


def class_is_archived(class_id: int, archived_class_ids: set[int]) -> bool:
    return class_id in archived_class_ids


def load_active_classes_by_ids(class_ids: list[int], db: Session) -> tuple[list[SchoolClass], set[int]]:
    archived_class_ids = read_archived_class_ids(db)
    classes = db.scalars(
        select(SchoolClass)
        .where(SchoolClass.id.in_(class_ids))
        .options(
            selectinload(SchoolClass.students),
            selectinload(SchoolClass.teacher_assignments),
            selectinload(SchoolClass.student_groups),
        )
    ).all()
    active_classes = [item for item in classes if not class_is_archived(item.id, archived_class_ids)]
    return active_classes, archived_class_ids


def ensure_assignable_class(class_id: int, db: Session) -> SchoolClass:
    school_class = db.get(SchoolClass, class_id)
    if school_class is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"班级不存在：{class_id}")
    if class_is_archived(class_id, read_archived_class_ids(db)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"班级已归档，不能再分配：{school_class.class_name}")
    return school_class


@router.get("/themes", response_model=ApiResponse)
def list_theme_presets(db: Session = Depends(get_db)) -> ApiResponse:
    system_settings = read_system_settings(db)
    return ApiResponse(
        data={
            "current_theme_code": system_settings.get("theme_code", "mango-splash"),
            "presets": theme_presets_payload(),
        }
    )


@router.post("/admin/promotions/preview", response_model=ApiResponse)
def preview_promotions(
    payload: AdminPromotionPreviewPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    source_classes, archived_class_ids = load_active_classes_by_ids(payload.source_class_ids, db)
    found_ids = {item.id for item in source_classes}
    missing_ids = [class_id for class_id in payload.source_class_ids if class_id not in found_ids]
    if missing_ids:
        blocked_archived_ids = [class_id for class_id in missing_ids if class_id in archived_class_ids]
        if blocked_archived_ids:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="所选班级中包含已归档班级，无法升班")
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"存在不存在或已归档的班级 ID: {missing_ids}")

    all_classes = db.scalars(select(SchoolClass)).all()
    active_by_grade_no = {
        (item.grade_no, item.class_no): item
        for item in all_classes
        if item.id not in archived_class_ids
    }
    preview = build_promotion_preview(
        source_classes,
        active_by_grade_no,
        grade_increment=payload.grade_increment,
    )
    return ApiResponse(data=preview)


@router.post("/admin/promotions/execute", response_model=ApiResponse)
def execute_promotions(
    payload: AdminPromotionExecutePayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    source_classes, archived_class_ids = load_active_classes_by_ids(payload.source_class_ids, db)
    source_by_id = {item.id: item for item in source_classes}
    missing_ids = [class_id for class_id in payload.source_class_ids if class_id not in source_by_id]
    if missing_ids:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"存在不存在或已归档的班级 ID: {missing_ids}")

    all_classes = db.scalars(select(SchoolClass)).all()
    active_by_grade_no = {
        (item.grade_no, item.class_no): item
        for item in all_classes
        if item.id not in archived_class_ids
    }
    preview = build_promotion_preview(
        source_classes,
        active_by_grade_no,
        grade_increment=payload.grade_increment,
    )

    now = datetime.now()
    archive_records = read_class_archive_records(db)
    source_student_ids_by_class = {
        item.id: [profile.user_id for profile in item.students]
        for item in source_classes
    }
    existing_class_names = {item.class_name for item in all_classes}
    created_target_count = 0
    reused_target_count = 0
    moved_student_count = 0
    archived_count = 0
    blocked_count = 0
    details: list[dict] = []

    for item in preview["items"]:
        source_class_id = int(item["source_class_id"])
        source = source_by_id[source_class_id]
        if item.get("status") != "ready":
            blocked_count += 1
            details.append(
                {
                    "source_class_id": source.id,
                    "source_class_name": source.class_name,
                    "status": "blocked",
                    "reason": item.get("reason") or "当前班级不满足升班条件",
                }
            )
            continue

        target_grade_no = int(item["target_grade_no"])
        target_class_no = int(item["target_class_no"])
        target = active_by_grade_no.get((target_grade_no, target_class_no))
        target_created = False
        if target is None:
            target = SchoolClass(
                grade_no=target_grade_no,
                class_no=target_class_no,
                class_name=make_class_display_name(target_grade_no, target_class_no),
                head_teacher_name=source.head_teacher_name,
                default_room_id=source.default_room_id,
            )
            db.add(target)
            db.flush()
            active_by_grade_no[(target_grade_no, target_class_no)] = target
            created_target_count += 1
            target_created = True
        else:
            reused_target_count += 1

        source_student_ids = source_student_ids_by_class.get(source.id, [])
        source_student_count = len(source_student_ids)
        if source_student_ids:
            profiles = db.scalars(
                select(StudentProfile).where(
                    StudentProfile.user_id.in_(source_student_ids),
                    StudentProfile.class_id == source.id,
                )
            ).all()
            for profile in profiles:
                profile.class_id = target.id
                profile.grade_no = target_grade_no
            moved_student_count += len(profiles)
            db.execute(
                delete(ClassSeatAssignment).where(
                    ClassSeatAssignment.class_id == source.id,
                    ClassSeatAssignment.student_user_id.in_(source_student_ids),
                )
            )

        db.execute(delete(StudentGroup).where(StudentGroup.class_id == source.id))

        if payload.copy_teacher_assignments:
            source_teacher_ids = {
                assignment.staff_user_id
                for assignment in source.teacher_assignments
            }
            target_teacher_ids = set(
                db.scalars(
                    select(TeacherClassAssignment.staff_user_id).where(
                        TeacherClassAssignment.class_id == target.id
                    )
                ).all()
            )
            for staff_user_id in source_teacher_ids:
                if staff_user_id in target_teacher_ids:
                    continue
                db.add(
                    TeacherClassAssignment(
                        staff_user_id=staff_user_id,
                        class_id=target.id,
                    )
                )

        db.execute(delete(TeacherClassAssignment).where(TeacherClassAssignment.class_id == source.id))

        if payload.archive_source_classes:
            source_name_before_archive = source.class_name
            archived_class_name = make_archived_class_name(source_name_before_archive, now, existing_class_names)
            source.class_name = archived_class_name
            existing_class_names.add(archived_class_name)
            archived_class_ids.add(source.id)
            archived_count += 1
            archive_records.insert(
                0,
                {
                    "source_class_id": source.id,
                    "source_class_name": source_name_before_archive,
                    "target_class_id": target.id,
                    "target_class_name": target.class_name,
                    "moved_student_count": source_student_count,
                    "grade_increment": payload.grade_increment,
                    "archived_at": now.isoformat(timespec="seconds"),
                },
            )

        details.append(
            {
                "source_class_id": source.id,
                "source_class_name": item["source_class_name"],
                "target_class_id": target.id,
                "target_class_name": target.class_name,
                "target_created": target_created,
                "moved_student_count": source_student_count,
                "status": "completed",
            }
        )

    write_archived_class_ids(archived_class_ids, db)
    write_class_archive_records(archive_records[:500], db)
    db.commit()
    return ApiResponse(
        message=(
            f"升班完成：迁移学生 {moved_student_count} 人，"
            f"新建班级 {created_target_count} 个，复用班级 {reused_target_count} 个，"
            f"归档班级 {archived_count} 个，阻塞 {blocked_count} 个。"
        ),
        data={
            **bootstrap_payload(db),
            "promotion_preview": preview,
            "promotion_result": {
                "created_target_count": created_target_count,
                "reused_target_count": reused_target_count,
                "moved_student_count": moved_student_count,
                "archived_count": archived_count,
                "blocked_count": blocked_count,
                "details": details,
            },
        },
    )


@router.post("/admin/classes", response_model=ApiResponse)
def create_class(
    payload: AdminClassPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if payload.default_room_id is not None and db.get(ComputerRoom, payload.default_room_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")

    school_class = SchoolClass(
        grade_no=payload.grade_no,
        class_no=payload.class_no,
        class_name=make_class_display_name(payload.grade_no, payload.class_no),
        head_teacher_name=(payload.head_teacher_name or "").strip() or None,
        default_room_id=payload.default_room_id,
    )
    db.add(school_class)
    db.commit()
    return ApiResponse(message="班级已创建", data=bootstrap_payload(db))


@router.post("/admin/classes/batch", response_model=ApiResponse)
def create_classes_batch(
    payload: AdminClassBatchPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    seen_names: set[str] = set()
    created_count = 0
    updated_count = 0
    skipped_count = 0
    details: list[dict] = []

    room_cache = {room.id: room for room in db.scalars(select(ComputerRoom)).all()}
    archived_class_ids = read_archived_class_ids(db)
    class_by_name = {
        item.class_name: item
        for item in db.scalars(select(SchoolClass)).all()
        if item.id not in archived_class_ids
    }

    for item in payload.items:
        class_name = f"{item.grade_no}{item.class_no:02d}班"
        if class_name in seen_names:
            skipped_count += 1
            details.append({"class_name": class_name, "status": "skipped", "reason": "批次内重复"})
            continue
        seen_names.add(class_name)

        if item.default_room_id is not None and item.default_room_id not in room_cache:
            skipped_count += 1
            details.append({"class_name": class_name, "status": "skipped", "reason": "默认机房不存在"})
            continue

        existing = class_by_name.get(class_name)
        if existing is not None:
            if payload.overwrite_existing:
                existing.head_teacher_name = (item.head_teacher_name or "").strip() or None
                existing.default_room_id = item.default_room_id
                updated_count += 1
                details.append({"class_name": class_name, "status": "updated", "reason": "已更新现有班级"})
            else:
                skipped_count += 1
                details.append({"class_name": class_name, "status": "skipped", "reason": "班级已存在"})
            continue

        school_class = SchoolClass(
            grade_no=item.grade_no,
            class_no=item.class_no,
            class_name=class_name,
            head_teacher_name=(item.head_teacher_name or "").strip() or None,
            default_room_id=item.default_room_id,
        )
        db.add(school_class)
        db.flush()
        class_by_name[class_name] = school_class
        created_count += 1
        details.append({"class_name": class_name, "status": "created", "reason": "已创建"})

    db.commit()
    return ApiResponse(
        message=f"批量班级处理完成：新增 {created_count}，更新 {updated_count}，跳过 {skipped_count}",
        data={
            **bootstrap_payload(db),
            "batch_result": {
                "created_count": created_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "items": details,
            },
        },
    )


@router.put("/admin/classes/{class_id}", response_model=ApiResponse)
def update_class(
    class_id: int,
    payload: AdminClassPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    archived_class_ids = read_archived_class_ids(db)
    school_class = db.get(SchoolClass, class_id)
    if school_class is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="班级不存在")
    if class_is_archived(class_id, archived_class_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="已归档班级不可编辑")
    if payload.default_room_id is not None and db.get(ComputerRoom, payload.default_room_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")

    school_class.grade_no = payload.grade_no
    school_class.class_no = payload.class_no
    school_class.class_name = make_class_display_name(payload.grade_no, payload.class_no)
    school_class.head_teacher_name = (payload.head_teacher_name or "").strip() or None
    school_class.default_room_id = payload.default_room_id
    db.commit()
    return ApiResponse(message="班级已更新", data=bootstrap_payload(db))


@router.delete("/admin/classes/{class_id}", response_model=ApiResponse)
def delete_class(
    class_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    school_class = db.scalar(
        select(SchoolClass)
        .where(SchoolClass.id == class_id)
        .options(
            selectinload(SchoolClass.students),
            selectinload(SchoolClass.classroom_sessions),
            selectinload(SchoolClass.teacher_assignments),
            selectinload(SchoolClass.seat_assignments),
        )
    )
    if school_class is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="班级不存在")

    archived_class_ids = read_archived_class_ids(db)
    if class_is_archived(class_id, archived_class_ids):
        archived_class_ids.discard(class_id)
        write_archived_class_ids(archived_class_ids, db)
        archive_records = [
            item
            for item in read_class_archive_records(db)
            if not (
                isinstance(item, dict)
                and int(item.get("source_class_id") or 0) == class_id
            )
        ]
        write_class_archive_records(archive_records, db)

    if school_class.students or school_class.classroom_sessions:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="班级下已有学生或课堂记录，不能直接删除")

    db.delete(school_class)
    db.commit()
    return ApiResponse(message="班级已删除", data=bootstrap_payload(db))


@router.post("/admin/teachers", response_model=ApiResponse)
def create_teacher(
    payload: AdminTeacherPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if db.scalar(select(User).where(User.username == payload.username.strip())) is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="教师账号已存在")
    if not payload.password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="新建教师必须提供初始密码")

    from app.core.security import hash_password

    user = User(
        username=payload.username.strip(),
        password_hash=hash_password(payload.password),
        display_name=payload.display_name.strip(),
        user_type="staff",
    )
    db.add(user)
    db.flush()
    db.add(
        StaffProfile(
            user_id=user.id,
            is_admin=payload.is_admin,
            title=(payload.title or "").strip() or None,
        )
    )
    for class_id in sorted(set(payload.class_ids)):
        ensure_assignable_class(class_id, db)
        db.add(TeacherClassAssignment(staff_user_id=user.id, class_id=class_id))

    db.commit()
    return ApiResponse(message="教师已创建", data=bootstrap_payload(db))


@router.put("/admin/teachers/{staff_user_id}", response_model=ApiResponse)
def update_teacher(
    staff_user_id: int,
    payload: AdminTeacherPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    user = db.scalar(
        select(User)
        .where(User.id == staff_user_id, User.user_type == "staff")
        .options(
            selectinload(User.staff_profile),
            selectinload(User.teacher_class_assignments),
        )
    )
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教师不存在")

    duplicate = db.scalar(
        select(User).where(User.username == payload.username.strip(), User.id != staff_user_id)
    )
    if duplicate is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="教师账号已存在")

    from app.core.security import hash_password

    user.username = payload.username.strip()
    user.display_name = payload.display_name.strip()
    if payload.password:
        user.password_hash = hash_password(payload.password)

    if user.staff_profile is None:
        user.staff_profile = StaffProfile(user_id=user.id, is_admin=payload.is_admin)
    user.staff_profile.is_admin = payload.is_admin
    user.staff_profile.title = (payload.title or "").strip() or None

    for assignment in list(user.teacher_class_assignments):
        db.delete(assignment)
    db.flush()

    for class_id in sorted(set(payload.class_ids)):
        ensure_assignable_class(class_id, db)
        db.add(TeacherClassAssignment(staff_user_id=user.id, class_id=class_id))

    db.commit()
    return ApiResponse(message="教师已更新", data=bootstrap_payload(db))


@router.delete("/admin/teachers/{staff_user_id}", response_model=ApiResponse)
def delete_teacher(
    staff_user_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    user = db.scalar(
        select(User)
        .where(User.id == staff_user_id, User.user_type == "staff")
        .options(
            selectinload(User.teacher_class_assignments),
            selectinload(User.staff_profile),
        )
    )
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="教师不存在")
    if user.staff_profile and user.staff_profile.is_admin:
        admin_count = db.scalar(
            select(func.count(User.id))
            .join(StaffProfile, StaffProfile.user_id == User.id)
            .where(StaffProfile.is_admin == True)  # noqa: E712
        )
        if int(admin_count or 0) <= 1:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="至少需要保留一个管理员账号")

    for assignment in list(user.teacher_class_assignments):
        db.delete(assignment)
    if user.staff_profile is not None:
        db.delete(user.staff_profile)
    db.delete(user)
    db.commit()
    return ApiResponse(message="教师已删除", data=bootstrap_payload(db))


@router.post("/admin/students/import", response_model=ApiResponse)
async def import_students(
    file: UploadFile | None = File(default=None),
    update_existing: bool = Form(default=False),
    default_password: str = Form(default="123456"),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    if file is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要导入的学生文件")
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件内容为空")

    normalized_password = default_password.strip()
    if len(normalized_password) < 6:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="默认密码长度不能少于 6 位")

    rows = load_import_rows(file.filename or "students.csv", file_bytes)
    preview = build_student_import_preview(
        rows,
        db,
        default_password=normalized_password,
        update_existing=update_existing,
    )
    db.commit()
    return ApiResponse(
        message=(
            f"学生导入完成：新增 {preview['created_count']}，"
            f"更新 {preview['updated_count']}，跳过 {preview['skipped_count']}"
        ),
        data={
            **bootstrap_payload(db),
            "import_result": preview,
        },
    )


@router.post("/admin/rooms", response_model=ApiResponse)
def create_room(
    payload: AdminRoomPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    room = ComputerRoom(
        name=payload.name.strip(),
        row_count=payload.row_count,
        col_count=payload.col_count,
        description=(payload.description or "").strip() or None,
    )
    db.add(room)
    db.flush()

    for seat in create_room_seats(room, payload.ip_prefix, payload.ip_start):
        db.add(seat)

    db.commit()
    return ApiResponse(message="机房已创建", data=bootstrap_payload(db))


@router.put("/admin/rooms/{room_id}", response_model=ApiResponse)
def update_room(
    room_id: int,
    payload: AdminRoomPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    room = db.get(ComputerRoom, room_id)
    if room is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")

    room.name = payload.name.strip()
    room.row_count = payload.row_count
    room.col_count = payload.col_count
    room.description = (payload.description or "").strip() or None
    db.commit()
    return ApiResponse(message="机房已更新", data=bootstrap_payload(db))


@router.post("/admin/rooms/{room_id}/seats/import", response_model=ApiResponse)
async def import_room_seats(
    room_id: int,
    file: UploadFile = File(...),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    room = db.scalar(
        select(ComputerRoom)
        .where(ComputerRoom.id == room_id)
        .options(selectinload(ComputerRoom.seats))
    )
    if room is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")

    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要导入的座位表文件")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件内容为空")

    payload = build_import_seat_payload(room, load_import_rows(file.filename, file_bytes))
    validate_room_seat_payload(room, payload, db)
    return ApiResponse(message="座位表已导入到当前草稿，请确认后保存", data=payload.model_dump())


@router.put("/admin/rooms/{room_id}/seats", response_model=ApiResponse)
def update_room_seats(
    room_id: int,
    payload: AdminSeatListPayload,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    room = db.scalar(
        select(ComputerRoom)
        .where(ComputerRoom.id == room_id)
        .options(selectinload(ComputerRoom.seats))
    )
    if room is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")

    validate_room_seat_payload(room, payload, db)

    existing_by_id = {seat.id: seat for seat in room.seats}
    seen_ids: set[int] = set()
    existing_updates: list[tuple[ComputerSeat, AdminSeatPayload]] = []
    new_items: list[AdminSeatPayload] = []

    for item in payload.seats:
        if item.id is not None and item.id in existing_by_id:
            seat = existing_by_id[item.id]
            seen_ids.add(seat.id)
            existing_updates.append((seat, item))
        else:
            new_items.append(item)

    for temp_index, (seat, _) in enumerate(existing_updates, start=1):
        seat.row_no = -temp_index
        seat.col_no = -temp_index
    if existing_updates:
        db.flush()

    for seat, item in existing_updates:
        seat.row_no = item.row_no
        seat.col_no = item.col_no
        seat.seat_label = item.seat_label.strip()
        seat.ip_address = item.ip_address.strip()
        seat.hostname = (item.hostname or "").strip() or None
        seat.is_enabled = item.is_enabled

    for item in new_items:
        db.add(
            ComputerSeat(
                room_id=room.id,
                row_no=item.row_no,
                col_no=item.col_no,
                seat_label=item.seat_label.strip(),
                ip_address=item.ip_address.strip(),
                hostname=(item.hostname or "").strip() or None,
                is_enabled=item.is_enabled,
            )
        )

    for seat in list(room.seats):
        if seat.id in seen_ids:
            continue
        db.delete(seat)

    room.row_count = payload.row_count
    room.col_count = payload.col_count
    db.commit()
    return ApiResponse(message="机房座位已更新", data=bootstrap_payload(db))


@router.delete("/admin/rooms/{room_id}", response_model=ApiResponse)
def delete_room(
    room_id: int,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> ApiResponse:
    room = db.scalar(
        select(ComputerRoom)
        .where(ComputerRoom.id == room_id)
        .options(
            selectinload(ComputerRoom.seats),
            selectinload(ComputerRoom.classes),
        )
    )
    if room is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="机房不存在")
    if room.classes:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="已有班级绑定该机房，不能直接删除")

    db.delete(room)
    db.commit()
    return ApiResponse(message="机房已删除", data=bootstrap_payload(db))

